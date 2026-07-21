"""
Sprint 3 -- Days 18-20: Peer Percentile Engine
================================================
Computes PERCENT_RANK for 10 metrics within each of 11 peer groups.
Generates radar charts and peer_comparison.xlsx.
"""

import sys, io, os

if __name__ == "__main__":
    sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8', errors='replace')

from pathlib import Path
import sqlite3
import pandas as pd
import numpy as np

try:
    from openpyxl.styles import PatternFill, Font
except ImportError:
    PatternFill = Font = None

try:
    import matplotlib
    matplotlib.use("Agg")
    import matplotlib.pyplot as plt
except ImportError:
    plt = None

PROJECT_ROOT = Path(__file__).resolve().parent.parent.parent
DB_PATH = PROJECT_ROOT / "data" / "db" / "nifty100.db"
OUTPUT_DIR = PROJECT_ROOT / "output"
REPORTS_DIR = PROJECT_ROOT / "reports" / "radar_charts"
REPORTS_DIR.mkdir(parents=True, exist_ok=True)

# 10 metrics to rank
RANK_METRICS = {
    "return_on_equity_pct": {"label": "ROE %", "higher_better": True},
    "roce_pct": {"label": "ROCE %", "higher_better": True},
    "net_profit_margin_pct": {"label": "NPM %", "higher_better": True},
    "debt_to_equity": {"label": "D/E", "higher_better": False},
    "free_cash_flow_cr": {"label": "FCF (Cr)", "higher_better": True},
    "pat_cagr_5yr": {"label": "PAT CAGR 5yr", "higher_better": True},
    "revenue_cagr_5yr": {"label": "Rev CAGR 5yr", "higher_better": True},
    "eps_cagr_5yr": {"label": "EPS CAGR 5yr", "higher_better": True},
    "interest_coverage": {"label": "ICR", "higher_better": True},
    "asset_turnover": {"label": "Asset Turnover", "higher_better": True},
}

# 8 radar axes
RADAR_METRICS = [
    "return_on_equity_pct", "roce_pct", "net_profit_margin_pct",
    "debt_to_equity", "free_cash_flow_cr", "pat_cagr_5yr",
    "revenue_cagr_5yr", "composite_quality_score"
]
RADAR_LABELS = ["ROE", "ROCE", "NPM", "D/E (inv)", "FCF", "PAT CAGR", "Rev CAGR", "Composite"]


def compute_percent_rank(group_df, metric, higher_better=True):
    """
    Compute PERCENT_RANK for a metric within a peer group.
    For D/E (lower is better), invert: 1 - PERCENT_RANK.
    """
    vals = group_df[metric].dropna()
    if len(vals) <= 1:
        return pd.Series(0.5, index=group_df.index)

    ranks = vals.rank(method="average", ascending=higher_better)
    pct_rank = (ranks - 1) / (len(ranks) - 1)
    return pct_rank.reindex(group_df.index)


def compute_all_peer_percentiles(conn, year="2024-03"):
    """
    Compute percentile ranks for all 11 peer groups across 10 metrics.
    Returns DataFrame for peer_percentiles table.
    """
    df_fr = pd.read_sql(f"SELECT * FROM financial_ratios WHERE year='{year}'", conn)
    df_pg = pd.read_sql("SELECT * FROM peer_groups", conn)
    df_comp = pd.read_sql("SELECT id as company_id, company_name FROM companies", conn)

    df = df_fr.merge(df_comp, on="company_id", how="left")

    results = []
    for group_name, group_companies in df_pg.groupby("peer_group_name"):
        company_ids = group_companies["company_id"].tolist()
        group_df = df[df["company_id"].isin(company_ids)].copy()

        if group_df.empty:
            continue

        for metric, meta in RANK_METRICS.items():
            if metric not in group_df.columns:
                continue

            pct_ranks = compute_percent_rank(group_df, metric, meta["higher_better"])

            for idx, row in group_df.iterrows():
                cid = row["company_id"]
                val = row.get(metric)
                pct = pct_ranks.get(idx)
                if pd.isna(pct):
                    pct = None
                else:
                    pct = round(pct, 4)

                results.append({
                    "company_id": cid,
                    "peer_group_name": group_name,
                    "metric": metric,
                    "value": val if not pd.isna(val) else None,
                    "percentile_rank": pct,
                    "year": year,
                })

    return pd.DataFrame(results)


def generate_radar_chart(company_id, company_data, peer_avg, output_path):
    """Generate a radar/polar chart for a company vs peer group average."""
    if plt is None:
        return

    num_vars = len(RADAR_LABELS)
    angles = np.linspace(0, 2 * np.pi, num_vars, endpoint=False).tolist()
    angles += angles[:1]

    # Normalise values to 0-100 scale for display
    def normalise_vals(vals, peer_vals):
        normed = []
        for i, (v, p) in enumerate(zip(vals, peer_vals)):
            if v is None or pd.isna(v):
                normed.append(0)
                continue
            max_val = max(abs(v), abs(p), 1) * 1.5
            # Invert D/E
            if RADAR_METRICS[i] == "debt_to_equity":
                normed.append(max(0, min(100, (1 - v / max_val) * 100)))
            else:
                normed.append(max(0, min(100, (v / max_val) * 100)))
        return normed

    company_vals = [company_data.get(m, 0) for m in RADAR_METRICS]
    peer_vals = [peer_avg.get(m, 0) for m in RADAR_METRICS]

    comp_normed = normalise_vals(company_vals, peer_vals) + [normalise_vals(company_vals, peer_vals)[0]]
    peer_normed = normalise_vals(peer_vals, peer_vals) + [normalise_vals(peer_vals, peer_vals)[0]]

    fig, ax = plt.subplots(figsize=(8, 8), subplot_kw=dict(polar=True))
    ax.fill(angles, comp_normed, alpha=0.25, color="#2196F3")
    ax.plot(angles, comp_normed, linewidth=2, color="#2196F3", label=company_id)
    ax.plot(angles, peer_normed, linewidth=2, color="#FF9800", linestyle="--", label="Peer Avg")
    ax.set_xticks(angles[:-1])
    ax.set_xticklabels(RADAR_LABELS, size=9)
    ax.set_ylim(0, 100)
    ax.set_title(f"{company_id} vs Peer Group", size=14, fontweight="bold", pad=20)
    ax.legend(loc="upper right", bbox_to_anchor=(1.3, 1.1))
    plt.tight_layout()
    plt.savefig(output_path, dpi=150, bbox_inches="tight")
    plt.close()


def generate_all_radar_charts(conn, year="2024-03"):
    """Generate radar charts for all companies in peer groups."""
    if plt is None:
        print("  matplotlib not available — skipping radar charts")
        return 0

    df_fr = pd.read_sql(f"SELECT * FROM financial_ratios WHERE year='{year}'", conn)
    df_pg = pd.read_sql("SELECT * FROM peer_groups", conn)

    count = 0
    for group_name, group_companies in df_pg.groupby("peer_group_name"):
        company_ids = group_companies["company_id"].tolist()
        group_df = df_fr[df_fr["company_id"].isin(company_ids)]

        if group_df.empty:
            continue

        # Compute peer group average
        peer_avg = {}
        for m in RADAR_METRICS:
            if m in group_df.columns:
                peer_avg[m] = group_df[m].mean()
            else:
                peer_avg[m] = 0

        for _, row in group_df.iterrows():
            cid = row["company_id"]
            company_data = {m: row.get(m, 0) for m in RADAR_METRICS}
            output_path = REPORTS_DIR / f"{cid}_radar.png"
            generate_radar_chart(cid, company_data, peer_avg, output_path)
            count += 1

    return count


def export_peer_comparison_xlsx(conn, year="2024-03"):
    """
    Generate peer_comparison.xlsx with 11 sheets, colour-coded percentile ranks.
    """
    df_fr = pd.read_sql(f"SELECT * FROM financial_ratios WHERE year='{year}'", conn)
    df_pg = pd.read_sql("SELECT * FROM peer_groups", conn)
    df_comp = pd.read_sql("SELECT id as company_id, company_name FROM companies", conn)
    df_pct = pd.read_sql(f"SELECT * FROM peer_percentiles WHERE year='{year}'", conn)

    xlsx_path = OUTPUT_DIR / "peer_comparison.xlsx"

    green = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid") if PatternFill else None
    yellow = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid") if PatternFill else None
    red = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid") if PatternFill else None
    gold = PatternFill(start_color="FFD700", end_color="FFD700", fill_type="solid") if PatternFill else None

    metric_cols = list(RANK_METRICS.keys())

    with pd.ExcelWriter(xlsx_path, engine="openpyxl") as writer:
        for group_name, group_companies in df_pg.groupby("peer_group_name"):
            company_ids = group_companies["company_id"].tolist()
            benchmarks = group_companies[group_companies["is_benchmark"] == 1]["company_id"].tolist()

            # Build sheet data
            group_fr = df_fr[df_fr["company_id"].isin(company_ids)].copy()
            group_fr = group_fr.merge(df_comp, on="company_id", how="left")

            if group_fr.empty:
                continue

            # Add percentile columns
            for metric in metric_cols:
                pct_col = f"{metric}_pctile"
                metric_pct = df_pct[
                    (df_pct["peer_group_name"] == group_name) &
                    (df_pct["metric"] == metric)
                ][["company_id", "percentile_rank"]]
                metric_pct = metric_pct.rename(columns={"percentile_rank": pct_col})
                group_fr = group_fr.merge(metric_pct, on="company_id", how="left")

            # Select columns for output
            output_cols = ["company_id", "company_name"]
            for m in metric_cols:
                output_cols.append(m)
                pct_col = f"{m}_pctile"
                if pct_col in group_fr.columns:
                    output_cols.append(pct_col)

            available_cols = [c for c in output_cols if c in group_fr.columns]
            sheet_df = group_fr[available_cols].copy()

            # Add median summary row
            median_row = {"company_id": "MEDIAN", "company_name": "Group Median"}
            for m in metric_cols:
                if m in sheet_df.columns:
                    median_row[m] = sheet_df[m].median()
            sheet_df = pd.concat([sheet_df, pd.DataFrame([median_row])], ignore_index=True)

            # Truncate sheet name to 31 chars
            sheet_name = group_name[:31]
            sheet_df.to_excel(writer, sheet_name=sheet_name, index=False)

            if PatternFill is None:
                continue

            # Colour-code percentile columns
            ws = writer.sheets[sheet_name]
            col_names = list(sheet_df.columns)

            for m in metric_cols:
                pct_col = f"{m}_pctile"
                if pct_col not in col_names:
                    continue
                col_idx = col_names.index(pct_col) + 1

                for row_idx in range(2, len(sheet_df) + 1):  # skip header, skip median
                    cell = ws.cell(row=row_idx, column=col_idx)
                    try:
                        val = float(cell.value) if cell.value is not None else None
                    except (ValueError, TypeError):
                        val = None
                    if val is None:
                        continue
                    if val >= 0.75:
                        cell.fill = green
                    elif val >= 0.25:
                        cell.fill = yellow
                    else:
                        cell.fill = red

            # Highlight benchmark row with gold
            for row_idx in range(2, len(sheet_df) + 2):
                cid_cell = ws.cell(row=row_idx, column=1)
                if cid_cell.value in benchmarks:
                    for col_idx in range(1, len(col_names) + 1):
                        ws.cell(row=row_idx, column=col_idx).fill = gold

    return xlsx_path


def run_peer_engine():
    """Main entry point: compute peer percentiles, radar charts, Excel report."""
    print("=" * 60)
    print("  Sprint 3 -- Peer Percentile Engine")
    print("=" * 60)

    conn = sqlite3.connect(DB_PATH)
    conn.execute("PRAGMA foreign_keys = ON;")
    year = "2024-03"

    # 1. Compute peer percentiles
    print("  Computing peer percentile rankings...")
    df_pct = compute_all_peer_percentiles(conn, year)
    print(f"  Computed {len(df_pct)} percentile records across {df_pct['peer_group_name'].nunique()} groups")

    # 2. Write to SQLite
    print("  Writing peer_percentiles to SQLite...")
    conn.execute("DROP TABLE IF EXISTS peer_percentiles;")
    conn.execute("""
        CREATE TABLE peer_percentiles (
            company_id TEXT NOT NULL,
            peer_group_name TEXT NOT NULL,
            metric TEXT NOT NULL,
            value REAL,
            percentile_rank REAL,
            year TEXT NOT NULL,
            PRIMARY KEY(company_id, peer_group_name, metric, year),
            FOREIGN KEY(company_id) REFERENCES companies(id) ON DELETE CASCADE
        )
    """)
    df_pct.to_sql("peer_percentiles", conn, if_exists="append", index=False)
    count = pd.read_sql("SELECT COUNT(*) as cnt FROM peer_percentiles", conn).iloc[0]["cnt"]
    print(f"  peer_percentiles table: {count} rows")

    # 3. Generate radar charts
    print("  Generating radar charts...")
    chart_count = generate_all_radar_charts(conn, year)
    print(f"  Generated {chart_count} radar chart PNGs in reports/radar_charts/")

    # 4. Export peer comparison Excel
    print("  Exporting peer_comparison.xlsx...")
    xlsx_path = export_peer_comparison_xlsx(conn, year)
    print(f"  Done: {xlsx_path}")

    conn.close()

    print("\n" + "=" * 60)
    print("  PEER ENGINE COMPLETED SUCCESSFULLY!")
    print("=" * 60)


if __name__ == "__main__":
    run_peer_engine()

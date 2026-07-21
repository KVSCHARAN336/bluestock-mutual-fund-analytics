"""
Sprint 3 -- Days 15-17: Financial Screener Engine
===================================================
Loads screener_config.yaml, applies threshold filters to financial_ratios,
supports 6 preset screeners + custom thresholds.
Handles: Financials sector D/E skip, Debt Free ICR pass, composite scoring.
Exports colour-coded screener_output.xlsx.
"""

import sys, io, os

if __name__ == "__main__":
    sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8', errors='replace')

from pathlib import Path
import sqlite3
import pandas as pd
import numpy as np
import yaml

try:
    from openpyxl.styles import PatternFill
except ImportError:
    PatternFill = None

PROJECT_ROOT = Path(__file__).resolve().parent.parent.parent
DB_PATH = PROJECT_ROOT / "data" / "db" / "nifty100.db"
CONFIG_PATH = PROJECT_ROOT / "config" / "screener_config.yaml"
OUTPUT_DIR = PROJECT_ROOT / "output"
OUTPUT_DIR.mkdir(parents=True, exist_ok=True)


def load_config():
    """Load screener_config.yaml."""
    with open(CONFIG_PATH, "r") as f:
        return yaml.safe_load(f)


def load_universe(conn, year="2024-03"):
    """
    Load the full screener universe: financial_ratios joined with
    market_cap and sectors for the given year.
    """
    df_fr = pd.read_sql(
        f"SELECT * FROM financial_ratios WHERE year='{year}'", conn)
    df_mc = pd.read_sql(
        f"SELECT company_id, pe_ratio, pb_ratio, dividend_yield_pct, market_cap_crore "
        f"FROM market_cap WHERE year='{year}'", conn)
    df_sec = pd.read_sql("SELECT company_id, broad_sector FROM sectors", conn)
    df_comp = pd.read_sql("SELECT id as company_id, company_name FROM companies", conn)

    # Merge
    df = df_fr.merge(df_mc, on="company_id", how="left")
    df = df.merge(df_sec, on="company_id", how="left")
    df = df.merge(df_comp, on="company_id", how="left")

    return df


def load_prev_year_de(conn, year="2024-03"):
    """Load previous year D/E for turnaround watch (declining D/E check)."""
    try:
        parts = year.split("-")
        prev_year = f"{int(parts[0]) - 1}-{parts[1]}"
        df = pd.read_sql(
            f"SELECT company_id, debt_to_equity as de_prev FROM financial_ratios WHERE year='{prev_year}'", conn)
        return df
    except Exception:
        return pd.DataFrame(columns=["company_id", "de_prev"])


def is_financial_sector(row):
    """Check if company is in Financials broad_sector."""
    sector = row.get("broad_sector", "")
    return isinstance(sector, str) and "financ" in sector.lower()


def apply_filters(df, filters, conn=None, year="2024-03"):
    """
    Apply threshold filters to the universe DataFrame.
    Returns filtered DataFrame.
    """
    mask = pd.Series(True, index=df.index)

    for fkey, fval in filters.items():
        if fkey == "roe_min":
            mask &= df["return_on_equity_pct"].fillna(-999) >= fval
        elif fkey == "de_max":
            # Skip Financials sector companies for D/E filter
            is_fin = df["broad_sector"].str.contains("Financ", case=False, na=False)
            de_ok = df["debt_to_equity"].fillna(999) <= fval
            mask &= (de_ok | is_fin)
        elif fkey == "fcf_min":
            mask &= df["free_cash_flow_cr"].fillna(-999) >= fval
        elif fkey == "revenue_cagr_5yr_min":
            mask &= df["revenue_cagr_5yr"].fillna(-999) >= fval
        elif fkey == "revenue_cagr_3yr_min":
            mask &= df["revenue_cagr_3yr"].fillna(-999) >= fval
        elif fkey == "pat_cagr_5yr_min":
            mask &= df["pat_cagr_5yr"].fillna(-999) >= fval
        elif fkey == "pe_max":
            mask &= df["pe_ratio"].fillna(999) <= fval
        elif fkey == "pb_max":
            mask &= df["pb_ratio"].fillna(999) <= fval
        elif fkey == "dividend_yield_min":
            mask &= df["dividend_yield_pct"].fillna(-999) >= fval
        elif fkey == "dividend_payout_max":
            mask &= df["dividend_payout_ratio_pct"].fillna(999) <= fval
        elif fkey == "icr_min":
            # Debt Free label = ICR infinity, always passes
            is_debt_free = df["icr_label"] == "Debt Free"
            icr_ok = df["interest_coverage"].fillna(-999) >= fval
            mask &= (icr_ok | is_debt_free)
        elif fkey == "sales_min":
            # Need P&L sales data
            if conn is not None:
                df_sales = pd.read_sql(
                    f"SELECT company_id, sales FROM profitandloss WHERE year='{year}'", conn)
                df = df.merge(df_sales, on="company_id", how="left", suffixes=("", "_pl"))
                mask &= df["sales"].fillna(0) >= fval
        elif fkey == "de_declining":
            if fval and conn is not None:
                df_prev = load_prev_year_de(conn, year)
                df = df.merge(df_prev, on="company_id", how="left")
                has_prev = df["de_prev"].notna()
                declining = df["debt_to_equity"] < df["de_prev"]
                mask &= (~has_prev | declining)

    result = df[mask].copy()
    result = result.sort_values("composite_quality_score", ascending=False)
    return result


def compute_winsorised_composite(df):
    """
    Compute composite quality score (0-100) with P10/P90 winsorisation.
    35% Profitability + 30% Cash Quality + 20% Growth + 15% Leverage.
    """
    def winsorise_and_scale(series):
        """Winsorise at P10/P90, then min-max scale to 0-100."""
        s = series.copy()
        p10 = s.quantile(0.10)
        p90 = s.quantile(0.90)
        s = s.clip(p10, p90)
        rng = p90 - p10
        if rng == 0:
            return pd.Series(50.0, index=series.index)
        return ((s - p10) / rng) * 100

    score = pd.Series(0.0, index=df.index)
    weights_used = 0.0

    # Profitability (35%)
    if "return_on_equity_pct" in df.columns and df["return_on_equity_pct"].notna().any():
        score += winsorise_and_scale(df["return_on_equity_pct"].fillna(0)) * 0.15
        weights_used += 0.15
    if "roce_pct" in df.columns and df["roce_pct"].notna().any():
        score += winsorise_and_scale(df["roce_pct"].fillna(0)) * 0.10
        weights_used += 0.10
    if "net_profit_margin_pct" in df.columns and df["net_profit_margin_pct"].notna().any():
        score += winsorise_and_scale(df["net_profit_margin_pct"].fillna(0)) * 0.10
        weights_used += 0.10

    # Cash Quality (30%)
    if "free_cash_flow_cr" in df.columns and df["free_cash_flow_cr"].notna().any():
        score += winsorise_and_scale(df["free_cash_flow_cr"].fillna(0)) * 0.15
        weights_used += 0.15
    if "cfo_quality_score" in df.columns and df["cfo_quality_score"].notna().any():
        score += winsorise_and_scale(df["cfo_quality_score"].fillna(0)) * 0.10
        weights_used += 0.10
    fcf_pos = (df["free_cash_flow_cr"].fillna(-1) > 0).astype(float) * 100
    score += fcf_pos * 0.05
    weights_used += 0.05

    # Growth (20%)
    if "revenue_cagr_5yr" in df.columns and df["revenue_cagr_5yr"].notna().any():
        score += winsorise_and_scale(df["revenue_cagr_5yr"].fillna(0)) * 0.10
        weights_used += 0.10
    if "pat_cagr_5yr" in df.columns and df["pat_cagr_5yr"].notna().any():
        score += winsorise_and_scale(df["pat_cagr_5yr"].fillna(0)) * 0.10
        weights_used += 0.10

    # Leverage (15%) -- inverted: lower D/E = higher score
    if "debt_to_equity" in df.columns and df["debt_to_equity"].notna().any():
        de_inv = 100 - winsorise_and_scale(df["debt_to_equity"].fillna(0))
        score += de_inv * 0.10
        weights_used += 0.10
    if "interest_coverage" in df.columns and df["interest_coverage"].notna().any():
        score += winsorise_and_scale(df["interest_coverage"].fillna(0)) * 0.05
        weights_used += 0.05

    if weights_used > 0:
        score = score / weights_used * 1.0

    return score.round(2)


def export_screener_xlsx(results_dict, output_path, config):
    """
    Export screener results to Excel with colour-coded cells.
    Green = meets threshold, Red = fails threshold.
    """
    kpi_columns = [
        "company_id", "company_name", "broad_sector",
        "return_on_equity_pct", "roce_pct", "net_profit_margin_pct",
        "debt_to_equity", "interest_coverage", "asset_turnover",
        "free_cash_flow_cr", "revenue_cagr_5yr", "pat_cagr_5yr",
        "eps_cagr_5yr", "cfo_quality_score", "capex_intensity_pct",
        "pe_ratio", "pb_ratio", "dividend_yield_pct",
        "dividend_payout_ratio_pct", "composite_quality_score",
    ]

    green_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid") if PatternFill else None
    red_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid") if PatternFill else None

    with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
        for preset_key, df in results_dict.items():
            # Select available columns
            cols = [c for c in kpi_columns if c in df.columns]
            sheet_df = df[cols].reset_index(drop=True)

            sheet_name = preset_key[:31]  # Excel 31-char limit
            sheet_df.to_excel(writer, sheet_name=sheet_name, index=False)

            if PatternFill is None:
                continue

            # Colour-code based on preset filters
            ws = writer.sheets[sheet_name]
            preset_config = config["presets"].get(preset_key, {})
            filters = preset_config.get("filters", {})

            # Map filter keys to column names
            filter_col_map = {
                "roe_min": ("return_on_equity_pct", "min"),
                "de_max": ("debt_to_equity", "max"),
                "fcf_min": ("free_cash_flow_cr", "min"),
                "revenue_cagr_5yr_min": ("revenue_cagr_5yr", "min"),
                "revenue_cagr_3yr_min": ("revenue_cagr_3yr", "min"),
                "pat_cagr_5yr_min": ("pat_cagr_5yr", "min"),
                "pe_max": ("pe_ratio", "max"),
                "pb_max": ("pb_ratio", "max"),
                "dividend_yield_min": ("dividend_yield_pct", "min"),
                "dividend_payout_max": ("dividend_payout_ratio_pct", "max"),
            }

            for fkey, fval in filters.items():
                if fkey not in filter_col_map:
                    continue
                col_name, direction = filter_col_map[fkey]
                if col_name not in cols:
                    continue
                col_idx = cols.index(col_name) + 1  # 1-based

                for row_idx in range(2, len(sheet_df) + 2):  # skip header
                    cell = ws.cell(row=row_idx, column=col_idx)
                    try:
                        val = float(cell.value) if cell.value is not None else None
                    except (ValueError, TypeError):
                        val = None
                    if val is None:
                        continue
                    if direction == "min":
                        cell.fill = green_fill if val >= fval else red_fill
                    else:
                        cell.fill = green_fill if val <= fval else red_fill


def run_screener():
    """Main entry point: run all 6 preset screeners and export results."""
    print("=" * 60)
    print("  Sprint 3 -- Financial Screener Engine")
    print("=" * 60)

    config = load_config()
    conn = sqlite3.connect(DB_PATH)
    conn.execute("PRAGMA foreign_keys = ON;")

    year = "2024-03"
    df_universe = load_universe(conn, year)
    print(f"  Universe loaded: {len(df_universe)} companies for {year}")

    # Recompute composite score with winsorisation
    df_universe["composite_quality_score"] = compute_winsorised_composite(df_universe)

    results = {}
    for preset_key, preset_cfg in config["presets"].items():
        name = preset_cfg["name"]
        filters = preset_cfg["filters"]
        filtered = apply_filters(df_universe.copy(), filters, conn, year)
        count = len(filtered)
        status = "OK" if 5 <= count <= 50 else "WARN"
        print(f"  {name:30s}: {count:>3} companies  [{status}]")
        results[preset_key] = filtered

    # Export Excel
    xlsx_path = OUTPUT_DIR / "screener_output.xlsx"
    print(f"\n  Exporting {xlsx_path.name}...")
    export_screener_xlsx(results, xlsx_path, config)

    conn.close()
    print(f"  Done: {xlsx_path}")
    return results


if __name__ == "__main__":
    run_screener()
